import csv, openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active

with open('hard.csv', encoding='utf8') as f:
    reader = csv.reader(f, delimiter=',')
    for item in reader:
        ws.append(item)

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, ws.max_row + 1):
	for j in range(1, ws.max_column + 1):
		cell = get_column_letter(i) + str(j)
		sheet[cell] = ws.cell(row=i, column=j).value


sheet.insert_rows(1, 1)
sheet.delete_rows(idx=4)
sheet['A1'] = ' '
sheet['B1'] = 'personal1 '
sheet['C1'] = 'personal2 '
sheet['D1'] = 'personal3 '
sheet['E1'] = 'personal4 '
sheet['F1'] = 'personal5 '
sheet['G1'] = 'personal6 '

wb.save('tets_xlsx.xlsx')